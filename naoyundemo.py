"""
Naoyun SDK Python Comprehensive Demo (naoyundemo.py)
Full-featured UI demo based on tkinter + matplotlib, covering all public SDK interfaces.
Depends on the NaoyunSdkApi core SDK in naoyunsdk.py.
"""
import asyncio
import time
import datetime
import os
from collections import deque
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox

# ========== Import from Core SDK ==========
from naoyunsdk import (
    NaoyunSdkApi, EarSide, EEGResultIndex, SignalQuality,
    BleDeviceInfo, DeviceStateEventArgs, DataReceivedEventArgs,
    DeviceStatusNotificationEventArgs, ServerAuthResultEventArgs,
    MentalStateDataEventArgs, SpectrumDataEventArgs,
    MATPLOTLIB_AVAILABLE, AIOHTTP_AVAILABLE, OPENPYXL_AVAILABLE
)

# ========== Import Matplotlib ==========
if MATPLOTLIB_AVAILABLE:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

import numpy as np


# ==================== Tkinter UI App ====================

class NaoyunDemoApp:
    def __init__(self, root: tk.Tk, app_id: str, app_secret: str):
        self.root = root
        self.app_id = app_id
        self.app_secret = app_secret
        self.sdk = NaoyunSdkApi()
        sdk_ver = self.sdk.GetSdkVersion()
        self.root.title(f"Naoyun SDK - Python UI Demo {sdk_ver}")
        # Default maximized window
        self.root.state('zoomed')
        self.root.minsize(1366, 768)

        self.base_font = ("Segoe UI", 16)
        self.root.option_add("*Font", self.base_font)

        style = ttk.Style()
        style.configure(".", font=self.base_font)
        style.configure("TLabelframe.Label", font=("Segoe UI", 16, "bold"))
        style.configure("TNotebook.Tab", font=self.base_font, padding=[15, 5])
        style.map("TNotebook.Tab",
                  foreground=[("selected", "#005A9E")],
                  background=[("selected", "#E1F0FF")],
                  font=[("selected", ("Segoe UI", 16, "bold"))])
        style.configure("TButton", font=self.base_font)
        self.discovered_devices = []
        self.log_counter = 0

        # Recording data area
        self.is_recording = False
        self.record_start_time = 0
        self.record_buffer_left = []
        self.record_buffer_right = []
        self.last_saved_filename = None

        # Real-time plot data area
        self.live_points_count = 1000
        self.live_left_data = deque([0]*self.live_points_count, maxlen=self.live_points_count)
        self.live_right_data = deque([0]*self.live_points_count, maxlen=self.live_points_count)
        self.is_live_plotting = False
        self.is_spectrum_running = False
        self.is_mental_state_running = False

        self.setup_ui()
        # SDK Auto-initialization (Server authentication is automatic during connection)
        self.sdk.Initialize(app_id=self.app_id, app_secret=self.app_secret,
                            is_domestic=(self.var_server_region.get() == "domestic"))
        self.log("[SDK] Auto-initialized. Server authentication will execute after BLE connection.")
        self.setup_callbacks()
        # Track all after callback IDs for cleanup
        self._after_ids = set()

        # Flag for program closing
        self._is_closing = False
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        if MATPLOTLIB_AVAILABLE:
            self.root.after(100, self.update_live_plot)

    def update_live_plot(self):
        # If closing or window destroyed, return
        if self._is_closing or not self.winfo_exists():
            return

        if self.is_live_plotting:
            try:
                l_data = list(self.live_left_data)
                r_data = list(self.live_right_data)
                self.line_live_l.set_ydata(l_data)
                self.line_live_r.set_ydata(r_data)
                self.ax_live_l.relim()
                self.ax_live_l.autoscale_view(scalex=False, scaley=True)
                self.ax_live_r.relim()
                self.ax_live_r.autoscale_view(scalex=False, scaley=True)
                self.canvas_live.draw_idle()
            except tk.TclError:
                # Plot destroyed, ignore error
                pass

        # Use tracked after ID
        after_id = self.root.after(100, self.update_live_plot)
        self._after_ids.add(after_id)

    def winfo_exists(self):
        """Check if window still exists"""
        try:
            return self.root.winfo_exists()
        except:
            return False

    def on_closing(self):
        """Mandatory cleanup when window closes"""
        if self._is_closing:
            return

        self._is_closing = True
        print("[System] Starting resource cleanup...")

        # 1. Cancel all after callbacks
        for after_id in list(self._after_ids):
            try:
                self.root.after_cancel(after_id)
            except:
                pass
        self._after_ids.clear()
        if hasattr(self, '_resize_after_id') and self._resize_after_id:
            try:
                self.root.after_cancel(self._resize_after_id)
            except:
                pass

        # 2. Stop live plotting
        self.is_live_plotting = False

        # 3. Create cleanup task
        asyncio.create_task(self._force_cleanup())

    async def _force_cleanup(self):
        """Force cleanup of all resources"""
        try:
            # Stop algorithm tasks
            if self.sdk._spectrum_task and not self.sdk._spectrum_task.done():
                self.sdk._spectrum_task.cancel()
                try:
                    await asyncio.wait_for(asyncio.shield(self.sdk._spectrum_task), timeout=1.0)
                except (asyncio.CancelledError, asyncio.TimeoutError):
                    pass

            # Disconnect BLE
            if self.sdk.IsConnected:
                try:
                    await asyncio.wait_for(self.sdk.DisconnectAsync(), timeout=2.0)
                except asyncio.TimeoutError:
                    print("[Warning] BLE disconnect timeout, forcing continuation")

            # Close Matplotlib plots
            if MATPLOTLIB_AVAILABLE:
                try:
                    plt.close('all')
                except:
                    pass

        except Exception as e:
            print(f"Cleanup error: {e}")

        finally:
            # Force stop event loop
            self._stop_event_loop()

    def _stop_event_loop(self):
        """Force stop all loops and exit"""
        try:
            self.root.quit()
            # Delay window destruction
            self.root.after(100, self._final_destroy)

        except Exception as e:
            print(f"Stop loop error: {e}")
            os._exit(0)

    def _final_destroy(self):
        """Final destruction"""
        try:
            self.root.destroy()
        except:
            pass
        finally:
            import sys
            sys.exit(0)

    def run_async(self, coro):
        asyncio.create_task(coro)

    def log(self, msg: str):
        now = datetime.datetime.now().strftime("[%H:%M:%S] ")
        self.txt_log.insert(tk.END, now + msg + "\n")
        self.txt_log.see(tk.END)

    def setup_ui(self):
        main_pane = tk.PanedWindow(self.root, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=4)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left Panel
        left_frame = ttk.Frame(main_pane, width=360)
        main_pane.add(left_frame, minsize=320)

        lf_dev = ttk.LabelFrame(left_frame, text="BLE Scanner")
        lf_dev.pack(fill=tk.BOTH, expand=True, pady=5)
        self.btn_scan = ttk.Button(lf_dev, text="Scan Devices (5s)", command=self.do_scan)
        self.btn_scan.pack(fill=tk.X, padx=5, pady=5)

        self.list_devices = tk.Listbox(lf_dev, height=6, font=self.base_font)
        self.list_devices.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.btn_connect = ttk.Button(lf_dev, text="Connect Selected", command=self.do_connect)
        self.btn_connect.pack(fill=tk.X, padx=5, pady=2)
        self.btn_disconnect = ttk.Button(lf_dev, text="Disconnect",
            command=lambda: self.run_async(self.sdk.DisconnectAsync()))
        self.btn_disconnect.pack(fill=tk.X, padx=5, pady=2)

        lf_server = ttk.LabelFrame(left_frame, text="Server Settings")
        lf_server.pack(fill=tk.X, pady=5)

        self.var_server_region = tk.StringVar(value="domestic")

        ttk.Radiobutton(lf_server, text="Domestic (CN)", variable=self.var_server_region, value="domestic").pack(anchor=tk.W, padx=5, pady=2)
        ttk.Radiobutton(lf_server, text="International", variable=self.var_server_region, value="international").pack(anchor=tk.W, padx=5, pady=2)
        ttk.Button(lf_server, text="Apply", command=self.apply_server_settings).pack(fill=tk.X, padx=5, pady=5)

        lf_status = ttk.LabelFrame(left_frame, text="Device Status")
        lf_status.pack(fill=tk.X, pady=10)

        self.status_labels = {}
        status_items =[
            ('mac', "MAC: --"),
            ('battery', "Battery: --"),
            ('worn', "Wearing: --"),
            ('noise', "Noise: --"),
            ('touch', "Touch: --"),
            ('autoplay', "AutoPlay: --")
        ]
        for key, text in status_items:
            lbl = ttk.Label(lf_status, text=text, font=('Segoe UI', 14), wraplength=300)
            lbl.pack(anchor=tk.W, padx=5, pady=1, fill=tk.X)
            self.status_labels[key] = lbl

        btn_frame = ttk.Frame(lf_status)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(btn_frame, text="Refresh Status",
                  command=lambda: self.run_async(self.do_refresh_status())).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        ttk.Button(btn_frame, text="Get MAC",
                  command=lambda: self.run_async(self.do_get_mac())).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.lbl_status = ttk.Label(left_frame, text="Status: Disconnected",
                                   foreground="gray", font=('Segoe UI', 17, 'bold'), wraplength=320)
        self.lbl_status.pack(anchor=tk.W, pady=5, fill=tk.X)

        # Middle Panel: Log area
        mid_frame = ttk.Frame(main_pane, width=500)
        main_pane.add(mid_frame, minsize=400)
        lf_log = ttk.LabelFrame(mid_frame, text="System Log")
        lf_log.pack(fill=tk.BOTH, expand=True, pady=5)

        self.txt_log = scrolledtext.ScrolledText(lf_log, font=("Consolas", 14), height=45)
        self.txt_log.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Right Panel
        right_frame = ttk.Frame(main_pane, width=1200)
        main_pane.add(right_frame, minsize=360)
        notebook = ttk.Notebook(right_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=5)

        def _fix_matplotlib_layout():
            try:
                for canvas_attr in ['canvas_live', 'canvas_latest', 'canvas_spectrum']:
                    canvas = getattr(self, canvas_attr, None)
                    if canvas:
                        widget = canvas.get_tk_widget()
                        if widget.winfo_viewable():
                            w, h = widget.winfo_width(), widget.winfo_height()
                            if w > 10 and h > 10:
                                widget.event_generate("<Configure>", width=w, height=h)
                                canvas.draw_idle()
            except Exception:
                pass

        def on_tab_changed(event):
            self.root.after(50, _fix_matplotlib_layout)

        notebook.bind("<<NotebookTabChanged>>", on_tab_changed)

        tab_basic = ttk.Frame(notebook)
        notebook.add(tab_basic, text="Basic / Control")
        self._build_basic_tab(tab_basic)

        tab_record = ttk.Frame(notebook)
        notebook.add(tab_record, text="Record & Replay")
        self._build_record_tab(tab_record)

        if MATPLOTLIB_AVAILABLE:
            tab_live = ttk.Frame(notebook)
            notebook.add(tab_live, text="Live EEG Graph")
            self._build_live_tab(tab_live)

        tab_latest = ttk.Frame(notebook)
        notebook.add(tab_latest, text="Latest EEG")
        self._build_latest_eeg_tab(tab_latest)

        if MATPLOTLIB_AVAILABLE:
            tab_spectrum = ttk.Frame(notebook)
            notebook.add(tab_spectrum, text="Spectrum")
            self._build_spectrum_tab(tab_spectrum)

        tab_mental = ttk.Frame(notebook)
        notebook.add(tab_mental, text="Mental State")
        self._build_mental_state_tab(tab_mental)

        self.lbl_bottom = ttk.Label(self.root, text="Waiting for connection...", relief=tk.SUNKEN,
                                   wraplength=800, anchor=tk.W)
        self.lbl_bottom.pack(side=tk.BOTTOM, fill=tk.X)
        
        def _update_bottom_wrap(event):
            if event.width > 40:
                self.lbl_bottom.config(wraplength=event.width - 20)
        self.lbl_bottom.bind('<Configure>', _update_bottom_wrap)

        self.root.update() 

        def apply_pane_ratio(event=None):
            try:
                pane_width = main_pane.winfo_width()
                if pane_width < 100:
                    return
                left_w = max(320, int(pane_width * 0.20))
                remaining = pane_width - left_w
                mid_w = int(remaining * 0.25)
                right_w = remaining - mid_w

                min_mid = 200
                min_right = 300
                if left_w + mid_w + min_right > pane_width:
                    left_w = max(280, pane_width - min_mid - min_right)
                    remaining = pane_width - left_w
                    mid_w = int(remaining * 0.25)
                if mid_w < min_mid:
                    mid_w = min_mid
                    left_w = min(left_w, pane_width - mid_w - min_right)

                main_pane.sash_place(0, left_w, 0)
                main_pane.sash_place(1, left_w + mid_w, 0)
            except Exception:
                pass

        apply_pane_ratio()
        self.root.after(300, _fix_matplotlib_layout)

        self._resize_after_id = None
        def on_main_resize(event):
            if event.widget is not self.root:
                return
            if self._resize_after_id:
                self.root.after_cancel(self._resize_after_id)
            self._resize_after_id = self.root.after(100, apply_pane_ratio)

        self.root.bind('<Configure>', on_main_resize)
        
    def _build_basic_tab(self, parent):
        lf1 = ttk.LabelFrame(parent, text="EEG Data Stream")
        lf1.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(lf1, text="Start Stream",
                  command=lambda: (self.log("[UI] Data stream opened"), self.run_async(self.sdk.SendStartDataAsync()))).pack(side=tk.LEFT, padx=5, pady=5, expand=True)
        ttk.Button(lf1, text="Stop Stream",
                  command=lambda: (self.log("[UI] Data stream closed"), self.run_async(self.sdk.SendStopDataAsync()))).pack(side=tk.LEFT, padx=5, pady=5, expand=True)

        lf3 = ttk.LabelFrame(parent, text="Noise Reduction")
        lf3.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(lf3, text="Normal (0)", command=lambda: self.run_async(self.do_set_noise(0))).pack(side=tk.LEFT, padx=2, pady=5, expand=True)
        ttk.Button(lf3, text="Noise Mode (1)", command=lambda: self.run_async(self.do_set_noise(1))).pack(side=tk.LEFT, padx=2, pady=5, expand=True)
        ttk.Button(lf3, text="Env Sound (2)", command=lambda: self.run_async(self.do_set_noise(2))).pack(side=tk.LEFT, padx=2, pady=5, expand=True)

        lf4 = ttk.LabelFrame(parent, text="Touch & Auto-Play")
        lf4.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(lf4, text="Touch ON", command=lambda: self.run_async(self.do_set_touch(True))).grid(row=0, column=0, padx=5, pady=5, sticky='we')
        ttk.Button(lf4, text="Touch OFF", command=lambda: self.run_async(self.do_set_touch(False))).grid(row=0, column=1, padx=5, pady=5, sticky='we')
        ttk.Button(lf4, text="AutoPlay ON", command=lambda: self.run_async(self.do_set_autoplay(True))).grid(row=1, column=0, padx=5, pady=5, sticky='we')
        ttk.Button(lf4, text="AutoPlay OFF", command=lambda: self.run_async(self.do_set_autoplay(False))).grid(row=1, column=1, padx=5, pady=5, sticky='we')
        lf4.columnconfigure(0, weight=1)
        lf4.columnconfigure(1, weight=1)

    def _build_record_tab(self, parent):
        lf = ttk.LabelFrame(parent, text="Automated 1-Minute Data Recorder")
        lf.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        info_lbl = ttk.Label(lf, text="Click START to capture EEG data. Click STOP to finish and save.",
                 justify=tk.CENTER, wraplength=360)
        info_lbl.pack(pady=10, fill=tk.X, padx=10)
        def _update_info_wrap(event):
            if event.width > 40:
                info_lbl.config(wraplength=event.width - 30)
        lf.bind('<Configure>', _update_info_wrap)
        self.lbl_record_time = ttk.Label(lf, text="Status: Ready", font=('Segoe UI', 18, 'bold'), wraplength=360, justify=tk.CENTER)
        self.lbl_record_time.pack(pady=5, fill=tk.X, padx=10)
        def _update_record_wrap(event):
            if event.width > 40:
                self.lbl_record_time.config(wraplength=event.width - 20)
        lf.bind('<Configure>', _update_record_wrap)
        self.progress_record = ttk.Progressbar(lf, orient=tk.HORIZONTAL, length=350, mode='determinate')
        self.progress_record.pack(pady=5)

        self.btn_start_record = ttk.Button(lf, text="START RECORDING", command=self.do_start_record)
        self.btn_start_record.pack(pady=10, ipadx=10, ipady=10)
        ttk.Separator(lf, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        self.btn_plot_saved = ttk.Button(lf, text="Visualize Last Saved", command=self.do_plot_last_saved)
        self.btn_plot_saved.pack(pady=5, ipadx=10, ipady=5)
        self.btn_plot_saved.config(state=tk.DISABLED)

        if OPENPYXL_AVAILABLE:
            self.btn_save_excel = ttk.Button(lf, text="Save Last Record to Excel", command=self.do_save_excel)
            self.btn_save_excel.pack(pady=5, ipadx=10, ipady=5)
            self.btn_save_excel.config(state=tk.DISABLED)

    def _build_live_tab(self, parent):
        control_frame = ttk.Frame(parent)
        control_frame.pack(fill=tk.X, pady=5)

        self.btn_toggle_live = ttk.Button(control_frame, text="Enable Live Plot", command=self.toggle_live_plot)
        self.btn_toggle_live.pack(side=tk.LEFT, padx=10)
        ttk.Label(control_frame, text="Note: Requires 'Start Stream' to be active.").pack(side=tk.LEFT, padx=10)

        self.fig_live, (self.ax_live_l, self.ax_live_r) = plt.subplots(2, 1, figsize=(5, 3.5), dpi=100, constrained_layout=True)

        x_data = list(range(self.live_points_count))
        self.line_live_l, = self.ax_live_l.plot(x_data, [0]*self.live_points_count, color='blue', linewidth=0.8)
        self.ax_live_l.set_title("Left Ear (uV)", fontsize=15)
        self.ax_live_l.grid(True, linestyle='--', alpha=0.6)

        self.line_live_r, = self.ax_live_r.plot(x_data, [0]*self.live_points_count, color='red', linewidth=0.8)
        self.ax_live_r.set_title("Right Ear (uV)", fontsize=15)
        self.ax_live_r.grid(True, linestyle='--', alpha=0.6)

        self.canvas_live = FigureCanvasTkAgg(self.fig_live, master=parent)
        self.canvas_live.draw()
        self.canvas_live.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def _build_latest_eeg_tab(self, parent):
        control = ttk.Frame(parent)
        control.pack(fill=tk.X, pady=5)

        ttk.Label(control, text="Seconds (1-60):").pack(side=tk.LEFT, padx=5)
        self.spin_latest_sec = ttk.Spinbox(control, from_=1, to=60, width=8, font=self.base_font)
        self.spin_latest_sec.set(5)
        self.spin_latest_sec.pack(side=tk.LEFT, padx=5)

        self.chk_filtered = tk.BooleanVar(value=False)
        ttk.Checkbutton(control, text="Use Filtered Data", variable=self.chk_filtered).pack(side=tk.LEFT, padx=10)

        ttk.Button(control, text="Fetch & Plot", command=self.do_fetch_latest_eeg).pack(side=tk.LEFT, padx=10)
        if OPENPYXL_AVAILABLE:
            ttk.Button(control, text="Save to Excel", command=self.do_save_latest_eeg_excel).pack(side=tk.LEFT, padx=10)

        if MATPLOTLIB_AVAILABLE:
            self.fig_latest, (self.ax_latest_l, self.ax_latest_r) = plt.subplots(2, 1, figsize=(5, 3.5), dpi=100, constrained_layout=True)
            self.ax_latest_l.set_title("Left Ear Latest EEG", fontsize=15)
            self.ax_latest_l.set_ylabel("uV")
            self.ax_latest_l.grid(True, linestyle='--', alpha=0.6)
            self.ax_latest_r.set_title("Right Ear Latest EEG", fontsize=15)
            self.ax_latest_r.set_ylabel("uV")
            self.ax_latest_r.set_xlabel("Sample Points")
            self.ax_latest_r.grid(True, linestyle='--', alpha=0.6)

            self.canvas_latest = FigureCanvasTkAgg(self.fig_latest, master=parent)
            self.canvas_latest.draw()
            self.canvas_latest.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        else:
            ttk.Label(parent, text="Matplotlib not available.").pack(pady=20)

    def _build_spectrum_tab(self, parent):
        control = ttk.Frame(parent)
        control.pack(fill=tk.X, pady=10)

        self.btn_toggle_spectrum = ttk.Button(control, text="Start Spectrum Task", command=self.toggle_spectrum)
        self.btn_toggle_spectrum.pack(side=tk.LEFT, padx=10)

        ttk.Label(control, text="Requires 'Start Stream' to be active.").pack(side=tk.LEFT, padx=5)

        info = ttk.Frame(parent)
        info.pack(fill=tk.X, pady=5)
        self.lbl_spectrum_quality = ttk.Label(info, text="Signal Quality: Waiting...", font=('Segoe UI', 14))
        self.lbl_spectrum_quality.pack(side=tk.LEFT, padx=10)
        self.lbl_spectrum_raw = ttk.Label(info, text="Raw: --", font=('Segoe UI', 12), foreground="gray")
        self.lbl_spectrum_raw.pack(side=tk.LEFT, padx=10)

        self.fig_spectrum, (self.ax_spectrum_l, self.ax_spectrum_r) = plt.subplots(1, 2, figsize=(6, 3.5), dpi=100)
        self.ax_spectrum_l.set_title("Left Ear Spectrum", fontsize=15)
        self.ax_spectrum_l.set_ylabel("Power")
        self.ax_spectrum_r.set_title("Right Ear Spectrum", fontsize=15)
        self.ax_spectrum_r.set_ylabel("Power")

        self.fig_spectrum.tight_layout()
        self.canvas_spectrum = FigureCanvasTkAgg(self.fig_spectrum, master=parent)
        self.canvas_spectrum.draw()
        self.canvas_spectrum.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def _build_mental_state_tab(self, parent):
        control = ttk.Frame(parent)
        control.pack(fill=tk.X, pady=10)

        self.btn_toggle_mental = ttk.Button(control, text="Start Mental State Task", command=self.toggle_mental_state)
        self.btn_toggle_mental.pack(side=tk.LEFT, padx=10)

        ttk.Label(control, text="Interval: 1.0s").pack(side=tk.LEFT, padx=5)

        frame_metrics = ttk.LabelFrame(parent, text="Real-time Metrics")
        frame_metrics.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.mental_labels = {}
        metrics = [
            ('focus', "Focus"),
            ('fatigue', "Fatigue"),
            ('relax', "Relax"),
            ('calm', "Calm"),
            ('stress', "Stress")
        ]
        for key, name in metrics:
            f = ttk.Frame(frame_metrics)
            f.pack(fill=tk.X, pady=5, padx=10)
            ttk.Label(f, text=f"{name}:", font=('Segoe UI', 16), width=12).pack(side=tk.LEFT)
            lbl = ttk.Label(f, text="--", font=('Segoe UI', 16, 'bold'))
            lbl.pack(side=tk.LEFT, padx=10)
            self.mental_labels[key] = lbl


    # ==================== Operational Methods & UI Response ====================

    def do_fetch_latest_eeg(self):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")
        if not MATPLOTLIB_AVAILABLE:
            return messagebox.showerror("Error", "Matplotlib not available.")

        try:
            sec = int(self.spin_latest_sec.get())
            if sec < 1 or sec > 60:
                raise ValueError
        except ValueError:
            return messagebox.showerror("Error", "Seconds must be 1-60")

        use_filtered = self.chk_filtered.get()
        left_data = self.sdk.GetLatestEegData(EarSide.Left, sec, use_filtered)
        right_data = self.sdk.GetLatestEegData(EarSide.Right, sec, use_filtered)

        self.ax_latest_l.clear()
        self.ax_latest_l.plot(left_data, color='blue', linewidth=0.5)
        self.ax_latest_l.set_title(f"Left Ear Latest EEG ({sec}s, {'Filtered' if use_filtered else 'Raw'})", fontsize=15)
        self.ax_latest_l.set_ylabel("uV")
        self.ax_latest_l.grid(True, linestyle='--', alpha=0.6)

        self.ax_latest_r.clear()
        self.ax_latest_r.plot(right_data, color='red', linewidth=0.5)
        self.ax_latest_r.set_title(f"Right Ear Latest EEG ({sec}s, {'Filtered' if use_filtered else 'Raw'})", fontsize=15)
        self.ax_latest_r.set_ylabel("uV")
        self.ax_latest_r.set_xlabel("Sample Points")
        self.ax_latest_r.grid(True, linestyle='--', alpha=0.6)

        self.canvas_latest.draw()
        self.log(f"Fetched latest {sec}s EEG data")

    def do_save_latest_eeg_excel(self):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")
        if not OPENPYXL_AVAILABLE:
            return messagebox.showerror("Error", "openpyxl not installed. Run: pip install openpyxl")
        try:
            sec = int(self.spin_latest_sec.get())
            if sec < 1 or sec > 60:
                raise ValueError
        except ValueError:
            return messagebox.showerror("Error", "Seconds must be 1-60")
        use_filtered = self.chk_filtered.get()
        left_data = self.sdk.GetLatestEegData(EarSide.Left, sec, use_filtered)
        right_data = self.sdk.GetLatestEegData(EarSide.Right, sec, use_filtered)
        if not left_data and not right_data:
            return messagebox.showerror("Error", "No data available. Please start stream first.")
        filename = f"Naoyun_EEG_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            saved = self.sdk.save_eeg_to_excel(
                left_data, right_data, filename=filename,
                metadata={
                    "device": self.sdk.get_current_status().get('device_name', 'unknown'),
                    "mac_address": self.sdk.MacAddress or "unknown",
                    "duration_sec": sec,
                    "filtered": use_filtered,
                    "saved_at": datetime.datetime.now().isoformat()
                }
            )
            self.log(f"[Excel] Saved: {saved}")
            messagebox.showinfo("Success", f"Excel saved:\n{saved}")
        except Exception as e:
            self.log(f"[Excel] Save failed: {e}")
            messagebox.showerror("Error", f"Excel save failed: {e}")

    def toggle_spectrum(self):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")

        if self.is_spectrum_running:
            self.sdk.StopSpectrumTask()
            self.is_spectrum_running = False
            self.btn_toggle_spectrum.config(text="Start Spectrum Task")
            self.log("Spectrum task stopped")
        else:
            success = self.sdk.StartSpectrumTask()
            if success:
                self.is_spectrum_running = True
                self.btn_toggle_spectrum.config(text="Stop Spectrum Task")
                self.log("Spectrum task started")
            else:
                messagebox.showerror("Error", "Failed to start spectrum task")

    def toggle_mental_state(self):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")

        if self.is_mental_state_running:
            self.sdk.StopMentalStateTask()
            self.is_mental_state_running = False
            self.btn_toggle_mental.config(text="Start Mental State Task")
            self.log("Mental state task stopped")
            for lbl in self.mental_labels.values():
                lbl.config(text="--")
        else:
            success = self.sdk.StartMentalStateTask(1.0)
            if success:
                self.is_mental_state_running = True
                self.btn_toggle_mental.config(text="Stop Mental State Task")
                self.log("Mental state task started (1.0s interval)")
            else:
                messagebox.showerror("Error", "Failed to start mental state task")

    async def do_refresh_status(self):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")
        self.log("[UI] Manually refreshing status...")
        success = await self.sdk.SendInitCommandAsync()
        if success:
            self.log("[UI] Status refresh successful")
            self.update_status_display()
        else:
            self.log("[UI] Status refresh failed")

    async def do_get_mac(self):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")
        self.log("[UI] Getting MAC address...")
        await self.sdk.GetMacAddressAsync()
        await asyncio.sleep(0.5)
        mac = self.sdk.MacAddress
        if mac:
            self.log(f"[UI] MAC Address: {mac}")
            self.status_labels['mac'].config(text=f"MAC: {mac}")
        else:
            self.log("[UI] Failed to get MAC address")

    async def do_set_noise(self, mode):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")
        success, verified = await self.sdk.SetNoiseReductionModeAsync(mode)
        if success and verified:
            self.log("[UI] Noise reduction mode set and verified")
        elif success:
            self.log("[UI] Noise reduction set but verification failed (device responded)")
        else:
            self.log("[UI] Failed to set noise reduction mode")
        self.update_status_display()

    async def do_set_touch(self, enabled: bool):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")
        success, verified = await self.sdk.SetTouchEnabledAsync(enabled)
        if success and verified:
            self.log(f"[UI] Touch set and verified: {'ON' if enabled else 'OFF'}")
        elif success:
            self.log(f"[UI] Touch set but verification failed: {'ON' if enabled else 'OFF'}")
        else:
            self.log("[UI] Failed to set touch")
        self.update_status_display()

    async def do_set_autoplay(self, enabled: bool):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect first!")
        success, verified = await self.sdk.SetAutoPlayEnabledAsync(enabled)
        if success and verified:
            self.log(f"[UI] AutoPlay set and verified: {'ON' if enabled else 'OFF'}")
        elif success:
            self.log(f"[UI] AutoPlay set but verification failed: {'ON' if enabled else 'OFF'}")
        else:
            self.log("[UI] Failed to set AutoPlay")
        self.update_status_display()

    def update_status_display(self):
        status = self.sdk.get_current_status()
        mac = self.sdk.MacAddress or "--"
        self.status_labels['mac'].config(text=f"MAC: {mac}")
        lb = status.get('left_battery', '--')
        rb = status.get('right_battery', '--')
        lb_str = "Not Connected" if lb == 0xFF else f"{lb}%"
        rb_str = "Not Connected" if rb == 0xFF else f"{rb}%"
        self.status_labels['battery'].config(text=f"Battery: Left {lb_str} Right {rb_str}")
        worn_str = f"L:{'Yes' if status.get('left_worn') else 'No'} R:{'Yes' if status.get('right_worn') else 'No'}"
        self.status_labels['worn'].config(text=f"Wearing: {worn_str}")
        noise_map = {0: "Off", 1: "ANC", 2: "Ambient"}
        noise_str = noise_map.get(status.get('noise_reduction'), f"Unknown({status.get('noise_reduction')})")
        self.status_labels['noise'].config(text=f"Noise: {noise_str}")
        self.status_labels['touch'].config(text=f"Touch: {'ON' if status.get('touch_enabled') else 'OFF'}")
        self.status_labels['autoplay'].config(text=f"AutoPlay: {'ON' if status.get('autoplay_enabled') else 'OFF'}")

    def do_scan(self):
        self.list_devices.delete(0, tk.END)
        self.discovered_devices.clear()
        self.log("Scanning devices...")
        self.btn_scan.config(state=tk.DISABLED)
        async def scan_task():
            await self.sdk.StartBleScanAsync(datetime.timedelta(seconds=5))
            self.log(f"Scan complete. Found {len(self.discovered_devices)} devices.")
            self.btn_scan.config(state=tk.NORMAL)
        self.run_async(scan_task())

    def do_connect(self):
        sel = self.list_devices.curselection()
        if not sel:
            return
        device = self.discovered_devices[sel[0]]
        self.lbl_status.config(text=f"Connecting to {device.Name}...", foreground="orange")
        self.run_async(self.sdk.ConnectBleAsync(device.Id))

    def apply_server_settings(self):
        if self.sdk.IsConnected:
            self.log("[UI] Please disconnect before switching server.")
            return
        is_domestic = self.var_server_region.get() == "domestic"
        self.sdk.Initialize(app_id=self.app_id, app_secret=self.app_secret,
                            is_domestic=is_domestic)
        region_str = "Domestic(CN)" if is_domestic else "International"
        self.log(f"[UI] Server config applied: {region_str}")

    def toggle_live_plot(self):
        self.is_live_plotting = not self.is_live_plotting
        if self.is_live_plotting:
            self.btn_toggle_live.config(text="Disable Live Plot")
        else:
            self.btn_toggle_live.config(text="Enable Live Plot")

    def do_start_record(self):
        if not self.sdk.IsConnected:
            return messagebox.showerror("Error", "Please connect to device first!")
        if self.is_recording:
            self.stop_and_save_record()
            return
        self.is_recording = True
        self.record_buffer_left.clear()
        self.record_buffer_right.clear()
        self.record_start_time = time.time()
        self.btn_start_record.config(text="STOP RECORDING")
        self.btn_plot_saved.config(state=tk.DISABLED)
        if OPENPYXL_AVAILABLE and hasattr(self, 'btn_save_excel'):
            self.btn_save_excel.config(state=tk.DISABLED)
        self.log("=== Started data recording (manual stop) ===")
        self.run_async(self.sdk.SendStartDataAsync())
        self.check_record_progress()

    def check_record_progress(self):
        if not self.is_recording:
            return
        elapsed = time.time() - self.record_start_time
        self.progress_record['value'] = (elapsed % 60) / 60.0 * 100
        self.lbl_record_time.config(text=f"Recording... {int(elapsed)} seconds")
        self.root.after(500, self.check_record_progress)

    def stop_and_save_record(self):
        self.is_recording = False
        self.progress_record['value'] = 100
        self.lbl_record_time.config(text="Saving... Please wait.")
        self.run_async(self.sdk.SendStopDataAsync())
        self.root.update()
        filename = f"Naoyun_EEG_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            saved_path = self.sdk.save_eeg_to_excel(
                self.record_buffer_left, self.record_buffer_right, filename=filename,
                metadata={"saved_at": datetime.datetime.now().isoformat()}
            )
            self.last_saved_filename = saved_path
            self.lbl_record_time.config(text=f"Saved:\n{saved_path}")
            self.log("Recording saved successfully")
            if MATPLOTLIB_AVAILABLE:
                self.btn_plot_saved.config(state=tk.NORMAL)
            if OPENPYXL_AVAILABLE and hasattr(self, 'btn_save_excel'):
                self.btn_save_excel.config(state=tk.NORMAL)
        except Exception as e:
            self.log(f"Save failed: {e}")
            messagebox.showerror("Error", f"Save failed: {e}")
        self.btn_start_record.config(text="START RECORDING")
        self.progress_record['value'] = 0

    def do_plot_last_saved(self):
        if not self.last_saved_filename:
            return messagebox.showerror("Error", "No saved file found.")
        left_data, right_data = [], []
        try:
            if self.last_saved_filename.endswith('.xlsx') and OPENPYXL_AVAILABLE:
                from openpyxl import load_workbook
                wb = load_workbook(self.last_saved_filename)
                for sheet_name, target_list in [("Left_EEG", left_data), ("Right_EEG", right_data)]:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        start = False
                        for row in ws.iter_rows(values_only=True):
                            if row and row[0] == sheet_name.replace("_EEG", "_uV"):
                                start = True
                                continue
                            if start and row and row[0] is not None:
                                try:
                                    target_list.append(float(row[0]))
                                except (ValueError, TypeError):
                                    pass
            else:
                base = self.last_saved_filename[:-4] if self.last_saved_filename.endswith('.csv') else self.last_saved_filename
                left_file = f"{base}_Left.csv"
                right_file = f"{base}_Right.csv"
                if os.path.exists(left_file) and os.path.exists(right_file):
                    for csv_file, target_list, label in [(left_file, left_data, "Left_uV"), (right_file, right_data, "Right_uV")]:
                        with open(csv_file, 'r') as f:
                            start = False
                            for line in f:
                                line = line.strip()
                                if line == label:
                                    start = True
                                    continue
                                if start and line:
                                    target_list.append(float(line))
                elif os.path.exists(self.last_saved_filename):
                    # Legacy combined CSV
                    with open(self.last_saved_filename, 'r') as f:
                        lines = f.readlines()[1:]
                        for line in lines:
                            parts = line.strip().split(',')
                            if len(parts) >= 2:
                                if parts[0]:
                                    left_data.append(float(parts[0]))
                                if parts[1]:
                                    right_data.append(float(parts[1]))
                else:
                    return messagebox.showerror("Error", "No saved file found.")
        except Exception as e:
            return messagebox.showerror("Read Error", str(e))

        top = tk.Toplevel(self.root)
        top.title(f"Post-Analysis Viewer: {self.last_saved_filename}")
        top.geometry("900x600")
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(7, 4.5), sharex=True, constrained_layout=True)
        ax1.plot(left_data, color='blue', linewidth=0.5)
        ax1.set_title(f"Left Ear (Total: {len(left_data)} points)", fontsize=15)
        ax1.set_ylabel("uV")
        ax1.grid(True, linestyle='--', alpha=0.5)
        ax2.plot(right_data, color='red', linewidth=0.5)
        ax2.set_title(f"Right Ear (Total: {len(right_data)} points)", fontsize=15)
        ax2.set_ylabel("uV")
        ax2.set_xlabel("Sample Points")
        ax2.grid(True, linestyle='--', alpha=0.5)
        canvas = FigureCanvasTkAgg(fig, master=top)
        canvas.draw()
        toolbar = NavigationToolbar2Tk(canvas, top)
        toolbar.update()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def do_save_excel(self):
        """Save the latest recording to Excel"""
        if not self.record_buffer_left and not self.record_buffer_right:
            return messagebox.showerror("Error", "No record data available. Please record first.")

        filename = f"Naoyun_EEG_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            saved = self.sdk.save_eeg_to_excel(
                self.record_buffer_left, self.record_buffer_right, filename=filename,
                metadata={"saved_at": datetime.datetime.now().isoformat()}
            )
            self.log(f"[Excel] Saved: {saved}")
            messagebox.showinfo("Success", f"Excel saved (2 sheets):\n{saved}")
        except Exception as e:
            self.log(f"[Excel] Save failed: {e}")
            messagebox.showerror("Error", f"Excel save failed: {e}")

    def setup_callbacks(self):
        self.sdk.on_log_message = self.log
        
        def on_raw(uuid, data_hex):
            pass
        self.sdk.on_raw_data_received = on_raw
        
        def on_device_discovered(info):
            if info.Id not in [d.Id for d in self.discovered_devices]:
                self.discovered_devices.append(info)
                self.list_devices.insert(tk.END, f"{info.Name}[{info.Id}]")
        self.sdk.on_ble_device_discovered = on_device_discovered

        def on_conn_state(args):
            color = "green" if args.IsConnected else "black"
            self.lbl_status.config(text=f"Status: {args.Message}", foreground=color)
            if args.IsConnected:
                self.log("Device connected")
        self.sdk.on_connection_state_changed = on_conn_state

        def on_status_notify(args):
            bat_l_str = "N/A" if args.LeftBattery == 0xFF else f"{args.LeftBattery}%"
            bat_r_str = "N/A" if args.RightBattery == 0xFF else f"{args.RightBattery}%"
            s = f"SW:{args.SoftwareVersion} HW:{args.HardwareVersion} BatL:{bat_l_str} BatR:{bat_r_str}"
            self.lbl_bottom.config(text=s)
            self.update_status_display()
        self.sdk.on_device_status_notification_received = on_status_notify
         
        def on_data(args):
            if self.is_recording:
                if args.EarSide == EarSide.Left:
                    self.record_buffer_left.extend(args.Data)
                elif args.EarSide == EarSide.Right:
                    self.record_buffer_right.extend(args.Data)

            if MATPLOTLIB_AVAILABLE:
                if args.EarSide == EarSide.Left:
                    self.live_left_data.extend(args.Data)
                elif args.EarSide == EarSide.Right:
                    self.live_right_data.extend(args.Data)
        self.sdk.on_data_received = on_data     

        def on_spectrum(args):
            if not hasattr(self, 'canvas_spectrum'):
                return
            quality_text = f"L: {args.LeftSignalQuality.name} | R: {args.RightSignalQuality.name}"
            self.lbl_spectrum_quality.config(text=f"Signal Quality: {quality_text}")

            labels = ['Delta', 'Theta', 'Alpha', 'Beta', 'Gamma']
            left_vals = [args.LeftDelta, args.LeftTheta, args.LeftAlpha, args.LeftBeta, args.LeftGamma]
            right_vals = [args.RightDelta, args.RightTheta, args.RightAlpha, args.RightBeta, args.RightGamma]

            self.ax_spectrum_l.clear()
            self.ax_spectrum_l.bar(labels, left_vals, color='blue', alpha=0.7)
            self.ax_spectrum_l.set_title("Left Ear Spectrum", fontsize=15)
            self.ax_spectrum_l.set_ylabel("Power")

            self.ax_spectrum_r.clear()
            self.ax_spectrum_r.bar(labels, right_vals, color='red', alpha=0.7)
            self.ax_spectrum_r.set_title("Right Ear Spectrum", fontsize=15)
            self.ax_spectrum_r.set_ylabel("Power")

            self.fig_spectrum.tight_layout()
            self.canvas_spectrum.draw()
        self.sdk.SpectrumDataReceived = on_spectrum

        def on_mental_state(args):
            self.mental_labels['focus'].config(text=f"{args.Focus:.1f}")
            self.mental_labels['fatigue'].config(text=f"{args.Fatigue:.1f}")
            self.mental_labels['relax'].config(text=f"{args.Relax:.1f}")
            self.mental_labels['calm'].config(text=f"{args.Calm:.1f}")
            self.mental_labels['stress'].config(text=f"{args.Stress:.1f}")
        self.sdk.MentalStateDataReceived = on_mental_state

        def on_auth(args):
            if args.IsSuccess:
                self.log(f"[Server Auth] Success, Permission: {args.Permission}")
            else:
                self.log(f"[Server Auth] Failed: {args.ErrorMessage}")
        self.sdk.ServerAuthCompleted = on_auth

async def app_main_loop(root: tk.Tk, app: 'NaoyunDemoApp'):
    """Main loop with exit detection"""
    try:
        while True:
            try:
                if app._is_closing or not app.winfo_exists():
                    break
                root.update()
                await asyncio.sleep(0.01)
            except tk.TclError:
                break
            except Exception as e:
                print(f"Main loop error: {e}")
                break
    except asyncio.CancelledError:
        pass
    finally:
        print("[System] Main loop exited")


if __name__ == "__main__":
    if hasattr(asyncio, 'WindowsSelectorEventLoopPolicy'):
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    root = tk.Tk()
    style = ttk.Style()
    if "vista" in style.theme_names():
        style.theme_use("vista")

    import sys
    if len(sys.argv) < 3:
        print("Usage: python naoyundemo.py <APP_ID> <APP_SECRET>")
        sys.exit(1)
    app_id = sys.argv[1]
    app_secret = sys.argv[2]
    app = NaoyunDemoApp(root, app_id, app_secret)

    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(app_main_loop(root, app))
    except KeyboardInterrupt:
        pass
    except RuntimeError as e:
        if "Event loop stopped before Future completed" not in str(e):
            print(f"Exception: {e}")
    except Exception as e:
        print(f"Exception: {e}")
    finally:
        if not app._is_closing:
            app.on_closing()
        import sys
        sys.exit(0)
